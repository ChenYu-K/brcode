from glob import glob
import csv
import numpy as np
from openpyxl import load_workbook
import matplotlib.pyplot as plt
 

filenames = glob('*.csv')
###
m=1;  #生データ書き込むためcellの列数，
row1 = 6  #計算結果を書き込むための行数
############################
#　エクセルシート（テンプレート）を読み込む
wb = load_workbook("template_mt.xlsx", data_only=True)  #エクセルシートを導入する
wb1 = wb.active           #シートを選択する

######
for filename in filenames:
  fname = filename.split('_')[-2]
  with open(filename,encoding="utf8",errors='ignore') as f:
    f_csv = csv.reader(f)
    next(f_csv)
    data = list(f_csv)
    a = np.array(data)
    load = a[1:,2]
    a = a[1:,:]         #文字列の次から
    x = range(np.shape(a)[0])
    y = len(data) - 1     #行数を数える
    case1 = np.empty([y,8],dtype = float)
  
  #########################
  for i in x:
    case1[i,0] = a[i,2]  #load
    case1[i,1] = a[i,3]  #disp
    if float(a[float(y/2),4]) > 0 :   #縦ひずみを判定する
      case1[i,2] = a[i,4]       #縦ひずみ1を定義する
      if float(a[float(y/2),5]) > 0 : #縦ひずみを判定する
        case1[i,3] = a[i,5]     #縦ひずみ2を定義する
        case1[i,4] = a[i,6]     #横ひずみ1を定義する
        case1[i,5] = a[i,7]     #横ひずみ2を定義する
      else :
        case1[i,4] = a[i,5]      #横ひずみ1を定義する
        case1[i,5] = a[i,7]    #横ひずみ2を定義する
        case1[i,3] = a[i,6]    #縦ひずみ2を定義する
    if i == y :
      break
    striantate = (case1[:,2]+case1[:,3])/2        #縦ひずみ平均値
    strianyoko = (case1[:,4]+case1[:,5])/2        #横ひずみ平均値

  #################
  
  A = wb1.cell(row1,6).value   #断面積を取得する
  case1[:,6] = case1[:,0] *1000/ A     #応力の算出
  
  #################降伏点および引張強度の算出
  sigma_y = 0
  i = 0
  while striantate[i] < 4000:
    sigma_y= max(sigma_y,case1[i,6])
    i = i+1
  sigma_u = max(case1[:,6])
  ##############ヤング率の算出
  i = 0; a1 = 0; a21 = 0; a22 = 0; k1 = 0; b1 = 0; n=0;pc = 0;
  for j in case1[:,6]:
    if j > 0.2*sigma_y and j < 0.7*sigma_y:
      a1 += (striantate[i] * case1[i,6])
      a21 += striantate[i]
      a22 += case1[i,6]
      b1 += striantate[i] ** 2
      pc += strianyoko[i]/striantate[i]
      n += 1
    i = i+1

  k1 = (n*a1 - a21*a22)/(n*b1-a21**2)     #yong's modulus
  b0 = a22/n - k1*a21/n            #intercept　切片
  pc = -pc/n                   #Poisson coefficient ポアソン比
  ####### R決定係数の求め
  i = 0; ssr = 0; sst = 0
  ymean = a22/n
  for j in case1[:,6]:
    if j > 0.2*sigma_y and j < 0.7*sigma_y:
      ssr += (case1[i,6] - (striantate[i]*k1+b0))**2
      sst += (case1[i,6] - ymean)**2
    i += 1
  R2 = 1 - ssr/sst        #R2 Coefficient of determination，決定係数
  ##########グラフを書く
  plt.plot(striantate,case1[:,6], label="stress-strain",color='black',linewidth=0.75)
  plt.xlabel("Strain")
  plt.ylabel("Stress [$N/mm^2$]")
  plt.ylim(0, 500)
  plt.xlim(0, 70000)
  plt.grid(color="k", linestyle=":") # メッシュ背景を点線に設定する
  plt.legend(loc=4,numpoints=1)            #凡例
  plt.savefig(fname+'-stress-strain.svg')    #svgとして保存
  plt.show()
  ###############データの書き込む
  wb1.cell(row1,10,sigma_y)
  wb1.cell(row1,11,sigma_u)
  wb1.cell(row1,12,k1)
  wb1.cell(row1,13,pc)

  for i in range(15,y) :
    j1 = 0
    for j in range(m,m+6) :
      wb1.cell(i,j,case1[i-15,j1]) #シートにi行j列にデータを書き込む
      j1 +=1
  
  m += 6
  row1 += 1

  wb.save("test.xlsx")     #保存する